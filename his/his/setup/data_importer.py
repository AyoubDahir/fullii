import frappe
from frappe.utils import getdate
from openpyxl import load_workbook

def create_rooms():
    try:
        file_path = "/home/hussein/frappe-bench/royal_decemebr.xlsx"

        # Load Excel file
        wb = load_workbook(file_path)
        sheet = wb.active

        # Read header row (first row)
        headers = [cell.value for cell in sheet[1]]

        # Parse all rows into dictionaries
        formatted_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))

            employee_id = row_dict.get("Employee")
            if not employee_id:
                continue

            # Extract shifts (all columns except these two)
            shifts = {
                key: value
                for key, value in row_dict.items()
                if key not in ["Employee", "Employee Name"]
            }

            formatted_data.append({
                "Employee": employee_id,
                "shifts": shifts
            })

        # Process and insert schedules
        for f_data in formatted_data:
            emp_id = f_data['Employee']
            emp_name = frappe.db.get_value("Employee", emp_id, "employee_name")

            for key, value in f_data['shifts'].items():

                if not value or not isinstance(value, str):
                    continue

                value = value.strip().upper()

                # Shift mapping
                shift_map = {
                    "D": "Day Shift",
                    "N": "Night Shift",
                    "DN": "Day and Night Shift",
                    "ND": "Night Day Shift",
                    "CANTEEN": "CANTEEN",
                    "OFF": "Free",
                    "OF": "Free",
                }

                shift = shift_map.get(value)
                if not shift:
                    continue

                # Convert key (column header) to day number
                try:
                    day_int = int(key)
                    shift_date = f"2025-12-{day_int:02d}"
                except:
                    continue

                # Create Doc
                sched_doc = frappe.get_doc({
                    "doctype": "Employee Schedulling",  # make sure this doctype exists!
                    "employee": emp_id,
                    "employee_name": emp_name,
                    "shift": shift,
                    "from_date": getdate(shift_date),
                    "to_date": getdate(shift_date),
                    "day": key,
                    "label": shift_date,
                    "month": "December",
                    "year": "2025"
                })

                sched_doc.insert()

        frappe.db.commit()
        print("✔ Schedule successfully imported!")

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Create Rooms Error")
        print(f"Error: {e}")



# import frappe
# import pandas as pd
# from frappe.utils import getdate

# def create_rooms():
#     df = pd.read_excel(r'/home/hussein/frappe-bench/royal_decemebr.xlsx')
#     df = pd.DataFrame(df)
#     data = df.to_dict(orient='records')
    
#     formatted_data = []
#     for item in data:
#         formatted_item = {'Employee': item['Employee']}
#         shifts = {key: value for key, value in item.items() if key not in ['Employee', 'Employee Name']}
#         formatted_item['shifts'] = shifts
#         formatted_data.append(formatted_item)

#     try:
#         for f_data in formatted_data:
#             emp_id = f_data['Employee']
#             emp_name = frappe.db.get_value("Employee", emp_id, "employee_name")
#             for key, value in f_data['shifts'].items():
#                 if not isinstance(value, str):
#                     continue
#                 value = value.strip().upper()
#                 shift = ""
#                 if value == "D":
#                     shift = "Day Shift"
#                 elif value == "N":
#                     shift = "Night Shift"
#                 elif value == "DN":
#                     shift = "Day and Night Shift"
#                 elif value == "ND":
#                     shift = "Night Day Shift"
#                 elif value == "CANTEEN":
#                     shift = "CANTEEN"
#                 elif value in ["OFF", "OF"]:
#                     shift = "Free"
                
#                 if not shift:
#                     continue

#                 try:
#                     day_int = int(key)
#                     shift_date = f"2025-12-{day_int:02d}"
#                 except ValueError:
#                     continue

#                 sched_doc = frappe.get_doc({
#                     "doctype": "Employee Schedulling",  # Double check spelling!
#                     "employee": emp_id,
#                     "employee_name": emp_name,
#                     "shift": shift,
#                     "from_date": getdate(shift_date),
#                     "to_date": getdate(shift_date),
#                     "day": key,
#                     "label": shift_date,
#                     "month": "December",
#                     "year": "2025"
#                 })
#                 sched_doc.insert()

#         frappe.db.commit()
#     except Exception as error:
#         frappe.log_error(frappe.get_traceback(), "Create Rooms Error")
#         print(error)
